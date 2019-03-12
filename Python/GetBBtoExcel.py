#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import unicode_literals

__author__  = 'Yuanze Wei'
__version__ = '0.2'
__doc__     = 'Module for get BB'

import matplotlib.pyplot as plt
import numpy as np
import matplotlib
import re
import sys
import os
import openpyxl as ol


def main():

    listWord = ["EnhancedEraseDisturbRE", "EnhancedEraseDisturbEP", "EraseDisturbEP", "EraseDisturbRE"]
    txtFiles = [name for name in os.listdir('./') if name.endswith('.txt')]
    for keyWord in listWord:
        outWb    = ol.Workbook()
        outWs    = outWb.create_sheet(index=0)
        rowCnt   = 4
        row      = 4
        col      = 4
        for textLogName in txtFiles:
            listDut = []
            starRow = rowCnt
            row     = starRow
            col     = 4
            keyCnt  = 0
            waferNum = re.split(r'_', textLogName)
            with open(textLogName, 'rt') as datalogFile:
                for line in datalogFile:
                    if re.match(r'%s[A-Za-z0-9_.\+\-\*\/\s]*INC_BDBLK' %keyWord, line):
                        splitSpace  = re.split(r'\s*[\s]\s*', line)
                        splitSpace0 = splitSpace[0]
                        splitSpace1 = splitSpace[1]
                        splitSpace4 = splitSpace[4]
                        splitSpace6 = splitSpace[6]
                        intDut      = int(splitSpace1.replace('DUT', ''))
                        if intDut in listDut:
                            keyCnt += 1
                            if keyCnt == 1:
                                rowCnt = row
                            listDut = []
                            row     = starRow
                            col     = 4 + keyCnt * 4
                            # print(row, col)
                        listDut.append(intDut)
                        outWs.cell(row, col+0).value = waferNum[2]
                        outWs.cell(row, col+1).value = splitSpace1
                        outWs.cell(row, col+2).value = splitSpace4
                        outWs.cell(row, col+3).value = splitSpace6
                        row += 1
                if keyCnt == 0:
                    rowCnt = row
        saveExcel = ".\\" + keyWord + ".xlsx"
        outWb.save(saveExcel)


if __name__ == "__main__":
    main()